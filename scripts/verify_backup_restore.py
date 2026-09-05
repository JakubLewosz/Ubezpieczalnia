"""Exercise a consistent native development backup and restore into an owned test DB.

Stop Django, Vite, Celery worker and beat before running. The source database and
media are read-only here. Only a freshly created randomly named restore database
is dropped. The synthetic backup and report stay under ignored .local/backups.
"""

import hashlib
import json
import os
import shutil
import socket
import subprocess
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import psycopg
from dotenv import dotenv_values
from local_services import binary
from psycopg import sql

ROOT = Path(__file__).resolve().parents[1]

API_PROBE = r'''
import hashlib, json, os
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
import django
django.setup()
from django.conf import settings
from django.contrib.auth import get_user_model
from documents.models import Document
from extraction.models import ApprovedRevision
from correspondence.models import Attachment, Message, ReadReceipt
from correspondence.sync_models import Mailbox
from openpyxl import load_workbook
from rest_framework.test import APIClient

revision = ApprovedRevision.objects.order_by("id").first()
assert revision is not None, "Brak zatwierdzonej rewizji do weryfikacji eksportu."
document = revision.document
preview = Path(settings.MEDIA_ROOT) / "previews" / str(document.pk) / "1.png"
assert preview.is_file(), "Brak podglądu dokumentu zatwierdzonej rewizji."
user = get_user_model().objects.filter(is_active=True).first()
assert user is not None, "Brak aktywnego konta demonstracji."
urls = [f"/api/documents/{document.pk}/original/", f"/api/documents/{document.pk}/pages/1/",
        f"/api/revisions/{revision.pk}/export/"]
anonymous = APIClient(enforce_csrf_checks=True)
for url in urls:
    assert anonymous.get(url).status_code == 403, "Anonimowe API nie odmówiło dostępu."
authenticated = APIClient(enforce_csrf_checks=True)
authenticated.force_login(user)
responses = [authenticated.get(url) for url in urls]
assert all(response.status_code == 200 for response in responses), "Błąd autoryzowanego pobrania po restore."
original = b"".join(responses[0].streaming_content)
png = b"".join(responses[1].streaming_content)
assert hashlib.sha256(original).hexdigest() == document.checksum
assert hashlib.sha256(png).hexdigest() == hashlib.sha256(preview.read_bytes()).hexdigest()
workbook = load_workbook(BytesIO(responses[2].content), data_only=False)
assert workbook.sheetnames == ["Informacje", "Dane"]
info = dict(workbook["Informacje"].iter_rows(values_only=True))
assert str(info["ID zatwierdzonej rewizji"]) == str(revision.pk)
assert info["SHA-256 dokumentu"] == revision.document_checksum
rows = list(workbook["Dane"].iter_rows(min_row=2))
assert len(rows) == len(revision.fields)
for row, field in zip(rows, revision.fields, strict=True):
    assert [row[0].value, row[1].value, row[2].value] == [field["group"], field["index"], field["code"]]
    cell = row[4]
    expected = field["value"]
    if expected is None:
        assert cell.value is None
    elif field["type"] == "date":
        assert cell.value.date().isoformat() == expected
    elif field["type"] in {"integer", "decimal"}:
        assert Decimal(str(cell.value)) == Decimal(expected)
    else:
        assert cell.value == expected and cell.data_type != "f"
    assert not any(c.data_type == "f" for c in row)
responses[0].close()
responses[1].close()
assert not Mailbox.objects.filter(kind="imap", enabled=True).exists(), "Importer kopii nie został wstrzymany."
assert os.environ.get("MAIL_SYNC_ENABLED") == "false"
message = Message.objects.exclude(raw_file="").order_by("id").first()
attachment = Attachment.objects.exclude(file="").filter(blocked_reason="").order_by("id").first()
assert message and attachment, "Brak wiadomości i dozwolonego załącznika do próby kopii poczty."
mail_urls = [f"/api/messages/{message.pk}/raw/", f"/api/mail-attachments/{attachment.pk}/download/"]
for url, expected in zip(mail_urls, [message.raw_sha256, attachment.checksum], strict=True):
    assert anonymous.get(url).status_code == 403
    response = authenticated.get(url)
    assert response.status_code == 200
    assert "attachment" in response.get("Content-Disposition", "")
    assert "no-store" in response.get("Cache-Control", "")
    assert hashlib.sha256(b"".join(response.streaming_content)).hexdigest() == expected
    response.close()
mail_counts = {"messages": Message.objects.count(), "attachments": Attachment.objects.count(),
               "personal_reads": ReadReceipt.objects.count(), "mailboxes": Mailbox.objects.count()}
print(json.dumps({"anonymous_denied": 3, "authenticated_downloads": 3,
    "revision_id": revision.pk, "export_fields_verified": len(rows), "mail": mail_counts,
    "mail_anonymous_denied": 2, "mail_authenticated_downloads": 2,
    "external_sync_paused": True}, ensure_ascii=False))
'''


def digest_file(path):
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def media_inventory(directory):
    result = {}
    for path in sorted(directory.rglob("*")):
        if path.is_symlink():
            raise RuntimeError("Magazyn demonstracji zawiera symlink; nie skopiowano go.")
        if path.is_file():
            result[path.relative_to(directory).as_posix()] = digest_file(path)
    return result


def database_snapshot(connection, media):
    tables = [row[0] for row in connection.execute(
        "SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename"
    )]
    counts = {table: connection.execute(sql.SQL("SELECT count(*) FROM {}").format(
        sql.Identifier("public", table))).fetchone()[0] for table in tables}
    originals = {}
    for document_id, filename, checksum in connection.execute(
        "SELECT id, file, checksum FROM documents_document ORDER BY id"
    ):
        path = (media / filename).resolve()
        if not path.is_relative_to(media.resolve()) or not path.is_file():
            raise RuntimeError("Oryginał nie istnieje w prawidłowym magazynie.")
        actual = digest_file(path)
        if actual != checksum:
            raise RuntimeError("Suma kontrolna oryginału różni się od metadanych bazy.")
        originals[str(document_id)] = actual
    revisions = list(connection.execute(
        "SELECT id, fields, document_checksum FROM extraction_approvedrevision ORDER BY id"
    ))
    revision_digest = hashlib.sha256(json.dumps(revisions, sort_keys=True, ensure_ascii=False).encode()).hexdigest()
    mail_hashes = {}
    mailboxes = []
    for table in tables:
        if table.startswith("correspondence_"):
            rows = [row[0] for row in connection.execute(
                sql.SQL("SELECT to_jsonb(t) FROM {} AS t ORDER BY id").format(sql.Identifier("public", table))
            )]
            mail_hashes[table] = hashlib.sha256(json.dumps(rows, sort_keys=True).encode()).hexdigest()
            if table == "correspondence_mailbox":
                mailboxes = [{key: row[key] for key in (
                    "id", "enabled", "uidvalidity", "boundary_uid", "discovered_uid", "version"
                )} for row in rows]
    for table, file_column, checksum_column in (
        ("correspondence_message", "raw_file", "raw_sha256"),
        ("correspondence_attachment", "file", "checksum"),
    ):
        if table not in tables:
            continue
        query = sql.SQL("SELECT {}, {} FROM {} WHERE {} <> ''").format(
            sql.Identifier(file_column), sql.Identifier(checksum_column), sql.Identifier(table),
            sql.Identifier(file_column),
        )
        for filename, checksum in connection.execute(query):
            path = (media / filename).resolve()
            if not path.is_relative_to(media.resolve()) or not path.is_file() or digest_file(path) != checksum:
                raise RuntimeError("Plik poczty nie istnieje w magazynie lub ma nieprawidłową sumę kontrolną.")
    return {"table_counts": counts, "originals": originals, "revision_count": len(revisions),
            "revision_fields_sha256": revision_digest, "mail_table_sha256": mail_hashes,
            "mailbox_cursors": mailboxes}


def checked_process(command, environment, error_file, output=None):
    with error_file.open("ab") as errors:
        result = subprocess.run(command, env=environment, stdout=output, stderr=errors, check=False)
    if result.returncode:
        raise RuntimeError(f"Narzędzie testu kopii zakończyło się błędem; szczegóły są w {error_file}.")


def main():
    os.umask(0o077)
    cfg = {**dotenv_values(ROOT / ".env"), **os.environ}
    if cfg.get("DJANGO_ENV") != "development":
        raise SystemExit("Test kopii wymaga jawnego DJANGO_ENV=development.")
    if cfg.get("POSTGRES_HOST", "127.0.0.1") not in {"127.0.0.1", "localhost"}:
        raise SystemExit("Test obsługuje wyłącznie natywny, lokalny PostgreSQL demonstracji.")
    for port in [8000, 5173]:
        with socket.socket() as probe:
            probe.settimeout(0.5)
            if probe.connect_ex(("127.0.0.1", port)) == 0:
                raise SystemExit("Najpierw zatrzymaj aplikację, worker i beat. Port 8000 lub 5173 jest aktywny.")
    source_name = cfg.get("POSTGRES_DB", "broker_office")
    connection_args = {"host": cfg.get("POSTGRES_HOST", "127.0.0.1"),
                       "port": cfg.get("POSTGRES_PORT", "54329"),
                       "user": cfg.get("POSTGRES_USER", "broker"),
                       "password": cfg.get("POSTGRES_PASSWORD", "")}
    media = Path(cfg.get("MEDIA_ROOT") or ROOT / ".local/media").resolve()
    if not media.is_dir():
        raise SystemExit("Brak magazynu plików demonstracji.")
    env = {**os.environ, "PGHOST": str(connection_args["host"]), "PGPORT": str(connection_args["port"]),
           "PGUSER": str(connection_args["user"]), "PGPASSWORD": str(connection_args["password"])}
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    token = uuid.uuid4().hex[:12]
    backup = ROOT / ".local/backups" / f"{stamp}-{token}"
    backup.mkdir(parents=True, mode=0o700, exist_ok=False)
    restore_name = f"broker_restore_test_{token}"
    created = False
    try:
        with psycopg.connect(dbname=source_name, **connection_args) as source:
            before = database_snapshot(source, media)
            if not before["originals"] or not before["revision_count"]:
                raise RuntimeError("Najpierw wykonaj demonstrację z dokumentem i zatwierdzoną rewizją.")
            version = source.execute("SHOW server_version").fetchone()[0]
            if not version.startswith("17."):
                raise RuntimeError("Ten test demonstracji wymaga PostgreSQL 17.")
        source_files = media_inventory(media)
        dump = backup / "database.dump"
        with dump.open("wb") as output:
            checked_process([binary("pg_dump"), "--dbname", source_name, "--format=custom", "--no-owner",
                             "--no-privileges"], env, backup / "tools.log", output)
        shutil.copytree(media, backup / "media")
        if media_inventory(backup / "media") != source_files:
            raise RuntimeError("Skopiowany magazyn różni się od źródła.")
        with psycopg.connect(dbname=source_name, **connection_args) as source:
            if database_snapshot(source, media) != before or media_inventory(media) != source_files:
                raise RuntimeError("Źródło zmieniło się podczas kopii. Zatrzymaj wszystkie procesy zapisujące dane.")
        with psycopg.connect(dbname="postgres", autocommit=True, **connection_args) as admin:
            admin.execute(sql.SQL("CREATE DATABASE {} TEMPLATE template0").format(sql.Identifier(restore_name)))
            created = True
        checked_process([binary("pg_restore"), "--dbname", restore_name, "--no-owner", "--no-privileges",
                         "--exit-on-error", str(dump)], env, backup / "tools.log")
        with psycopg.connect(dbname=restore_name, **connection_args) as restored:
            after = database_snapshot(restored, backup / "media")
        if before != after:
            raise RuntimeError("Odtworzona baza różni się od źródła w licznikach, oryginałach lub rewizjach.")
        # Verify the exact restore first. Then pause ONLY the newly created clone,
        # retaining business states, personal reads, historical UIDs and cursors.
        with psycopg.connect(dbname=restore_name, **connection_args) as restored:
            if "correspondence_mailbox" in after["table_counts"]:
                restored.execute("""UPDATE correspondence_mailbox
                    SET enabled=false, state='disabled', error_code='restored_paused',
                        error_message='Kopia po odtworzeniu. Administrator musi sprawdzić stan i jawnie wznowić import.',
                        lease_token=NULL, lease_expires=NULL, queued_until=NULL, next_attempt_at=NULL,
                        version=version+1 WHERE kind='imap'""")
        probe_env = {**cfg, "POSTGRES_DB": restore_name, "MEDIA_ROOT": str(backup / "media"),
                     "MAIL_SYNC_ENABLED": "false",
                     "DJANGO_ALLOWED_HOSTS": "testserver,localhost,127.0.0.1",
                     "PYTHONPATH": str(ROOT / "backend")}
        probe_env = {str(key): str(value) for key, value in probe_env.items() if value is not None}
        result = subprocess.run([sys.executable, "-c", API_PROBE], cwd=ROOT, env=probe_env,
                                text=True, capture_output=True, check=False)
        if result.returncode:
            (backup / "api-probe.log").write_text(result.stderr, encoding="utf-8")
            raise RuntimeError(f"Test API odtworzonej kopii nie powiódł się; szczegóły w {backup / 'api-probe.log'}.")
        report = {"notice": "DANE TESTOWE - spójna kopia i próba odtworzenia", "created_at_utc": stamp,
                  "postgres_version": version, "source_database": source_name, "source_snapshot": before,
                  "database_dump_sha256": digest_file(dump), "media_hashes": source_files,
                  "restored_database_compared": True, "restored_external_sync_paused": True,
                  "api_probe": json.loads(result.stdout)}
        (backup / "manifest.json").write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"DANE TESTOWE: kopia i odtworzenie zweryfikowane. Raport: {backup / 'manifest.json'}")
        print(f"Porównano {len(before['table_counts'])} tabel, {len(before['originals'])} oryginałów, "
              f"{before['revision_count']} rewizji i {len(source_files)} plików magazynu.")
        print("API: anonimowy dostęp odrzucony; oryginał, PNG i historyczny XLSX zweryfikowane po zalogowaniu.")
    finally:
        if created:
            # This name is generated in this invocation and never supplied by the caller.
            with psycopg.connect(dbname="postgres", autocommit=True, **connection_args) as admin:
                admin.execute(sql.SQL("DROP DATABASE {} WITH (FORCE)").format(sql.Identifier(restore_name)))


if __name__ == "__main__":
    main()
