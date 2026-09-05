from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from exports.profile import build_workbook


def revision_fixture():
    fields = []
    for index, (value, field_type) in enumerate([
        ("0000123", "text"), ("=1+1", "text"), ("+cmd", "text"), ("-1+2", "text"),
        ("@SUM(1)", "text"), (" \t=HYPERLINK(\"https://demo.invalid\")", "text"),
        ("75000.25", "decimal"), ("2026-10-01", "date"), ("2024", "integer"),
        ("Zażółć gęślą jaźń DANE TESTOWE", "text"), (None, "decimal"),
    ]):
        fields.append({"group": "participants", "index": index, "code": f"test_{index}", "label": "DANE TESTOWE",
                       "value": value, "type": field_type, "unit": "PLN" if field_type == "decimal" else "",
                       "page": 1, "manual": index % 2 == 0, "absent": value is None})
    return SimpleNamespace(pk=7, document_id=3, document_name="=DANE TESTOWE.pdf", document_checksum="a" * 64,
                           number=2, profile="broker_motor_application_v0", fields=fields,
                           created_at=datetime(2026, 9, 1, tzinfo=timezone.utc))


def test_workbook_contains_real_dates_numbers_leading_zeros_and_never_formulas():
    revision = revision_fixture()
    payload = build_workbook(revision)
    workbook = load_workbook(BytesIO(payload))
    assert workbook.sheetnames == ["Informacje", "Dane"]
    sheet = workbook["Dane"]
    for row, field in enumerate(revision.fields, start=2):
        cell = sheet.cell(row, 5)
        assert cell.data_type != "f"
        if field["type"] == "text":
            assert cell.value == field["value"] and cell.data_type == "s"
        elif field["type"] == "decimal" and field["value"] is not None:
            assert cell.value == pytest.approx(float(field["value"])) and cell.data_type == "n"
        elif field["type"] == "date":
            assert cell.value == datetime(2026, 10, 1)
    assert sheet["E2"].value == "0000123"
    assert workbook["Informacje"]["B3"].value == "=DANE TESTOWE.pdf"
    assert workbook["Informacje"]["B3"].data_type == "s"
    with ZipFile(BytesIO(payload)) as archive:
        for name in archive.namelist():
            if name.startswith("xl/worksheets/"):
                assert b"<f>" not in archive.read(name)


def test_repeated_export_has_identical_snapshot_data():
    first = load_workbook(BytesIO(build_workbook(revision_fixture())))
    second = load_workbook(BytesIO(build_workbook(revision_fixture())))
    for name in first.sheetnames:
        assert list(first[name].values) == list(second[name].values)
