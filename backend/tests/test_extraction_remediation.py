"""A02/A03/A05/A06 regression via real Django/DRF and PostgreSQL."""
import copy
import hashlib
import json
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from django.core.files.base import ContentFile
from openpyxl import load_workbook

from documents.models import Document
from exports.profile import build_workbook
from exports.text import ExportValidationError
from extraction.acquisition import acquire_document
from extraction.engine import BrokerMotorEngine, PageText
from extraction.models import ApprovedRevision, EngineResult, ExtractionJob, ReviewDraft
from extraction.numbered import blank_profile
from extraction.validation import draft_warnings, warning_digest

FIXTURES = Path(__file__).resolve().parents[2] / "fixtures/remediation"
EXPECTED = json.loads((FIXTURES / "expected.json").read_text())
OCR_OBSERVATIONS = json.loads((FIXTURES / "ocr_observations.json").read_text())


def values(result):
    return {f"{f['group']}.{f['index']}.{f['code']}": f["value"] for f in result["fields"]}


@pytest.mark.parametrize("name", ["numbered", "variant", "holdout"])
def test_numbered_plain_text_and_text_pdf(name, settings, tmp_path, monkeypatch):
    settings.MEDIA_ROOT = tmp_path
    monkeypatch.setattr("extraction.acquisition._ocr", lambda _: pytest.fail("Text PDF must not use OCR"))
    document = SimpleNamespace(pk=1, file=SimpleNamespace(path=str(FIXTURES / f"{name}.pdf")), mime_type="application/pdf")
    for pages in [[PageText(1, "text", (FIXTURES / f"{name}.txt").read_text())], acquire_document(document)]:
        result = BrokerMotorEngine().extract(pages)
        assert result["profile"] == "broker_motor_application_v1"
        actual = values(result)
        for key, expected in EXPECTED[name].items():
            assert actual[key] == expected, key
        assert len({f["group_id"] for f in result["fields"] if f["group"] == "participants"}) == 1
        for field in result["fields"]:
            if field["value"] is not None:
                page = pages[field["page"] - 1]
                assert field["source"] in page.text, field


@pytest.mark.ocr
@pytest.mark.parametrize("name,methods", [("numbered_scan.pdf", ["ocr"]), ("numbered.png", ["ocr"]), ("numbered.jpg", ["ocr"]), ("numbered_mixed.pdf", ["text", "ocr"])])
def test_numbered_actual_ocr_images_and_mixed(name, methods, settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path
    document = SimpleNamespace(pk=1, file=SimpleNamespace(path=str(FIXTURES / name)), mime_type="image/png" if name.endswith("png") else "image/jpeg" if name.endswith("jpg") else "application/pdf")
    pages = acquire_document(document)
    assert [p.method for p in pages] == methods
    result = BrokerMotorEngine().extract(pages)
    assert result["profile"] == "broker_motor_application_v1"
    actual = values(result)
    for key, expected in EXPECTED["numbered"].items():
        if actual[key] != expected:
            # Measured real OCR mistakes remain explicit; never infer Pin -> PLN
            # or remove extra letter O from a registration number.
            allowed = OCR_OBSERVATIONS["accepted_readings"]
            assert actual[key] in allowed.get(key, []), (name, key, actual[key])
            field = next(f for f in result["fields"] if f"{f['group']}.{f['index']}.{f['code']}" == key)
            assert field["method"] == "ocr" and field["source"] and field["warnings"]
            if key.endswith("insured_sum"):
                assert field["unit_conflict"]
    assert any(f["warnings"] for f in result["fields"] if f["method"] == "ocr")


@pytest.mark.parametrize("prefix", ["To nie jest ", "> ", '"', "Wiadomość klienta: "])
def test_negated_and_quoted_header_does_not_establish_profile(prefix):
    text = (FIXTURES / "numbered.txt").read_text().replace("Wniosek brokerski nr", prefix + "Wniosek brokerski nr")
    assert BrokerMotorEngine().extract([PageText(1, "text", text)])["profile"] is None


@pytest.mark.parametrize("prefix", ["Temat: Wniosek komunikacyjny\n", "Polisa komunikacyjna\n", "Wniosek ubezpieczenia domu\n"])
def test_mail_policy_property_do_not_become_application_from_cited_layout(prefix):
    assert BrokerMotorEngine().extract([PageText(1, "text", prefix + (FIXTURES / "numbered.txt").read_text())])["profile"] is None


def test_ambiguous_vehicle_names_and_missing_person_do_not_get_guessed():
    text = (FIXTURES / "numbered.txt").read_text().replace("MarkaTestowa ModelTestowy", "Nieznana Marka Nieznany Model").replace("Anna Demonstracyjna,", "??? ,")
    result = values(BrokerMotorEngine().extract([PageText(1, "text", text)]))
    assert result["vehicle.0.make"] is result["vehicle.0.model"] is result["participants.0.name"] is None
    assert result["coverage.0.premium"] is None


@pytest.fixture
def review_document(user, customer):
    payload = (FIXTURES / "numbered.pdf").read_bytes()
    return Document.objects.create(client=customer, author=user, file=ContentFile(payload, name="DANE TESTOWE.pdf"),
                                   original_name="DANE TESTOWE.pdf", mime_type="application/pdf", size=len(payload), checksum=hashlib.sha256(payload).hexdigest(), page_count=1)


@pytest.fixture
def review_draft(review_document, user):
    parsed = BrokerMotorEngine().extract([PageText(1, "text", (FIXTURES / "numbered.txt").read_text())])
    job = ExtractionJob.objects.create(document=review_document, requested_by=user, status="succeeded")
    result = EngineResult.objects.create(job=job, **parsed)
    return ReviewDraft.objects.create(document=review_document, engine_result=result, fields=result.fields, profile=result.profile)


def approve(api, doc, draft, **extra):
    return api.post(f"/api/documents/{doc.pk}/approve/", {
        "version": draft["version"], "warning_digest": draft["warning_digest"], "confirm_warnings": True,
        "note": "DANE TESTOWE: sprawdzono źródło i zachowano wskazane braki.", **extra,
    }, format="json")


@pytest.mark.django_db
def test_add_missing_person_and_scope_remove_conflict_approve_export_history(api, review_draft, review_document):
    base = f"/api/documents/{review_document.pk}"
    draft = api.get(base + "/review/").data["draft"]
    original = approve(api, review_document, draft)
    assert original.status_code == 201, original.data
    snapshot = copy.deepcopy(original.data["fields"])
    added = api.post(base + "/review/groups/", {"version": draft["version"], "group": "participants"}, format="json")
    assert added.status_code == 200, added.data
    assert api.post(base + "/review/groups/", {"version": draft["version"], "group": "participants"}, format="json").status_code == 409
    draft = added.data
    fields = copy.deepcopy(draft["fields"])
    new = [f for f in fields if f["group"] == "participants" and f["index"] == 1]
    assert all(f["manual"] and f["method"] == "manual" and f["updated_by"] and f["updated_at"] and not f["source"] and f["page"] is None for f in new)
    for field in new:
        field["value"] = "Bruno DANE TESTOWE" if field["code"] == "name" else "policyholder,insured" if field["code"] == "role" else None
    saved = api.patch(base + "/review/", {"version": draft["version"], "fields": fields}, format="json")
    assert saved.status_code == 200, saved.data
    draft = api.post(base + "/review/groups/", {"version": saved.data["version"], "group": "coverage_items"}, format="json").data
    item = next(f for f in draft["fields"] if f["group"] == "coverage_items" and f["index"] == 3)
    deleted_id = item["group_id"]
    deleted = api.delete(base + "/review/groups/", {"version": draft["version"], "group_id": deleted_id}, format="json")
    assert deleted.status_code == 200
    draft = api.post(base + "/review/groups/", {"version": deleted.data["version"], "group": "coverage_items"}, format="json").data
    item = next(f for f in draft["fields"] if f["group"] == "coverage_items" and f["index"] == 4)
    assert item["group_id"] != deleted_id
    assert all(f["group_id"] != deleted_id for f in draft["fields"])
    for field in draft["fields"]:
        if field["group"] == "coverage_items" and field["index"] == 4:
            field["value"] = {"requested_scope": "AC", "insured_sum": "12345.50", "variant": "DANE TESTOWE"}[field["code"]]
    saved = api.patch(base + "/review/", {"version": draft["version"], "fields": draft["fields"]}, format="json")
    final = approve(api, review_document, saved.data)
    assert final.status_code == 201, final.data
    payload = api.get(f"/api/revisions/{final.data['id']}/export/")
    assert payload.status_code == 200
    rows = list(load_workbook(BytesIO(payload.content))["Dane"].values)
    assert sum(row[2] == "name" and row[4] in {"Anna Demonstracyjna", "Bruno DANE TESTOWE"} for row in rows[1:]) == 2
    assert any(row[2] == "insured_sum" and row[4] == 12345.5 for row in rows[1:])
    assert api.get(f"/api/revisions/{original.data['id']}/").data["fields"] == snapshot
    assert api.delete(base + "/review/groups/", {"version": saved.data["version"] - 1, "group_id": new[0]["group_id"]}, format="json").status_code == 409


@pytest.mark.django_db
@pytest.mark.parametrize("status", ["failed", "succeeded"])
def test_manual_rescue_has_no_fake_result_and_can_export(api, review_document, user, status):
    job = ExtractionJob.objects.create(document=review_document, requested_by=user, status=status)
    if status == "succeeded":
        EngineResult.objects.create(job=job, profile=None, fields=[], warnings=["Brak profilu automatycznego odczytu"])
    response = api.post(f"/api/documents/{review_document.pk}/review/manual/", {}, format="json")
    assert response.status_code == 201, response.data
    assert response.data["origin"] == "manual"
    assert response.data["profile"] == "broker_motor_application_v1"
    assert not EngineResult.objects.filter(job=job, profile__isnull=False).exists()
    assert all(f["manual"] and f["method"] == "manual" for f in response.data["fields"])
    assert api.post(f"/api/documents/{review_document.pk}/review/manual/", {}, format="json").status_code == 409
    approved = approve(api, review_document, response.data)
    assert approved.status_code == 201, approved.data
    assert api.get(f"/api/revisions/{approved.data['id']}/export/").status_code == 200


@pytest.mark.django_db
def test_current_validation_and_acknowledgement_bound_to_version_and_warnings(api, review_draft, review_document):
    base = f"/api/documents/{review_document.pk}"
    before = api.get(base + "/review/").data["draft"]
    assert api.post(base + "/approve/", {"version": before["version"]}, format="json").status_code == 400
    fields = copy.deepcopy(before["fields"])
    for field in fields:
        if field["code"] == "vin":
            field["value"] = "BAD-VIN"
        if field["code"] == "end_date":
            field["value"] = "2026-01-01"
        if field["code"] == "email":
            field["value"] = "nie-email"
    saved = api.patch(base + "/review/", {"version": before["version"], "fields": fields}, format="json")
    assert saved.status_code == 200
    draft = saved.data
    assert {w["code"] for w in draft["warnings"]} >= {"invalid_vin", "date_order", "invalid_email"}
    assert approve(api, review_document, draft, note="").status_code == 400
    assert approve(api, review_document, draft, warning_digest=before["warning_digest"]).status_code == 400
    accepted = approve(api, review_document, draft)
    assert accepted.status_code == 201
    assert accepted.data["warning_confirmation"]["version"] == draft["version"]
    for field in draft["fields"]:
        if field["code"] == "vin":
            field["value"] = "TEST1234567890123"
        if field["code"] == "end_date":
            field["value"] = "2027-09-03"
        if field["code"] == "email":
            field["value"] = "anna@broker-demo.invalid"
    fixed = api.patch(base + "/review/", {"version": draft["version"], "fields": draft["fields"]}, format="json")
    assert fixed.status_code == 200
    assert not {"invalid_vin", "date_order", "invalid_email"} & {w["code"] for w in fixed.data["warnings"]}
    assert approve(api, review_document, draft).status_code == 409


@pytest.mark.django_db
@pytest.mark.parametrize("value", ["2026-02-30", "zła data"])
def test_invalid_date_can_be_saved_and_explicitly_approved_as_text(api, review_draft, review_document, value):
    fields = copy.deepcopy(review_draft.fields)
    next(f for f in fields if f["code"] == "end_date")["value"] = value
    saved = api.patch(f"/api/documents/{review_document.pk}/review/", {"version": 1, "fields": fields}, format="json")
    assert saved.status_code == 200, saved.data
    assert any(w["code"] == "invalid_type_value" and w["requires_note"] for w in saved.data["warnings"])
    approved = approve(api, review_document, saved.data)
    assert approved.status_code == 201, approved.data
    payload = api.get(f"/api/revisions/{approved.data['id']}/export/")
    rows = list(load_workbook(BytesIO(payload.content))["Dane"].values)
    assert any(row[2] == "end_date" and row[4] == value for row in rows)


@pytest.mark.django_db
@pytest.mark.parametrize("value", ["DANE\u0001TESTOWE", "\u000b", "\ufffe", "A" * 32768], ids=["control1", "control11", "noncharacter", "too_long"])
def test_invalid_xlsx_text_rejected_in_current_draft_and_old_revision_controlled(api, review_draft, review_document, user, value):
    fields = copy.deepcopy(review_draft.fields)
    fields[0]["value"] = value
    saved = api.patch(f"/api/documents/{review_document.pk}/review/", {"version": 1, "fields": fields}, format="json")
    assert saved.status_code == 400 and ("U+" in str(saved.data) or "32767" in str(saved.data))
    revision = ApprovedRevision.objects.create(document=review_document, engine_result=review_draft.engine_result, number=1, draft_version=1, fields=fields,
        profile=review_draft.profile, document_name=review_document.original_name, document_checksum=review_document.checksum, author=user)
    response = api.get(f"/api/revisions/{revision.pk}/export/")
    assert response.status_code == 400 and "korektę" in str(response.data)
    revision.refresh_from_db()
    assert revision.fields[0]["value"] == value


@pytest.mark.django_db
@pytest.mark.parametrize("metadata", ["document_name", "profile", "label", "unit"])
def test_invalid_historical_metadata_is_checked(api, review_draft, review_document, user, metadata):
    data = dict(document=review_document, engine_result=review_draft.engine_result, number=1, draft_version=1, fields=copy.deepcopy(review_draft.fields),
                profile=review_draft.profile, document_name="DANE TESTOWE.pdf", document_checksum=review_document.checksum, author=user)
    if metadata in {"label", "unit"}:
        data["fields"][0][metadata] = "DANE\u0001TESTOWE"
    else:
        data[metadata] = "DANE\u0001TESTOWE"
    revision = ApprovedRevision.objects.create(**data)
    with pytest.raises(ExportValidationError):
        build_workbook(revision)
    assert api.get(f"/api/revisions/{revision.pk}/export/").status_code == 400


@pytest.mark.django_db
@pytest.mark.parametrize("value", ["\n\tZażółć gęślą jaźń", "0000123", "=SUM(1,2)", "+pole", "-pole", "@pole", " \t=1+1", "A" * 32767], ids=["unicode", "zeros", "equals", "plus", "minus", "at", "whitespace", "max_length"])
def test_valid_xlsx_text_roundtrips_without_formula_or_truncation(api, review_draft, review_document, value):
    fields = copy.deepcopy(review_draft.fields)
    fields[0]["value"] = value
    saved = api.patch(f"/api/documents/{review_document.pk}/review/", {"version": 1, "fields": fields}, format="json")
    assert saved.status_code == 200, saved.data
    approved = approve(api, review_document, saved.data)
    assert approved.status_code == 201, approved.data
    payload = api.get(f"/api/revisions/{approved.data['id']}/export/")
    cell = load_workbook(BytesIO(payload.content))["Dane"]["E2"]
    assert cell.value == value and cell.data_type == "s"


@pytest.mark.django_db
def test_group_schema_role_spoof_and_reset_never_reuse_deleted_identity(api, review_draft, review_document):
    base = f"/api/documents/{review_document.pk}"
    assert api.post(base + "/review/groups/", {"version": 1, "group": "arbitrary", "fields": []}, format="json").status_code == 400
    fields = copy.deepcopy(review_draft.fields)
    next(f for f in fields if f["code"] == "role")["value"] = "administrator"
    assert api.patch(base + "/review/", {"version": 1, "fields": fields}, format="json").status_code == 400
    group_id = next(f for f in review_draft.fields if f["group"] == "participants")["group_id"]
    deleted = api.delete(base + "/review/groups/", {"version": 1, "group_id": group_id}, format="json")
    assert deleted.status_code == 200
    restored = api.post(base + "/review/reset/", {"version": deleted.data["version"]}, format="json")
    assert restored.status_code == 200
    assert all(f["group_id"] != group_id for f in restored.data["fields"])
    assert next(f for f in restored.data["fields"] if f["group"] == "participants")["index"] > 0


def test_source_conflict_stays_important_until_manual_correction():
    fields = blank_profile()
    fields[0].update(value=None, source_conflict=True, warnings=["Sprzeczne wartości źródła."])
    warnings = draft_warnings(fields)
    assert any(w["code"] == "source_conflict" and w["requires_note"] for w in warnings)
    assert warning_digest(fields)


@pytest.mark.django_db
def test_source_conflict_survives_unrelated_edit_and_clears_on_manual_field_change(api, review_draft, review_document):
    fields = copy.deepcopy(review_draft.fields)
    fields[0].update(value=None, source_conflict=True, warnings=["Sprzeczne wartości źródła."], method="text")
    review_draft.fields = fields
    review_draft.save()
    fields[1]["value"] = "2026-08-28"
    first = api.patch(f"/api/documents/{review_document.pk}/review/", {"version": 1, "fields": fields}, format="json")
    assert first.status_code == 200
    assert any(w["code"] == "source_conflict" and w["requires_note"] for w in first.data["warnings"])
    fields = first.data["fields"]
    fields[0]["value"] = "DANE TESTOWE wybrana wartość"
    second = api.patch(f"/api/documents/{review_document.pk}/review/", {"version": first.data["version"], "fields": fields}, format="json")
    assert second.status_code == 200
    assert not any(w["code"] == "source_conflict" for w in second.data["warnings"])


@pytest.mark.django_db
def test_approval_rejects_invalid_stored_roles_and_units_cannot_be_changed(api, review_draft, review_document):
    fields = copy.deepcopy(review_draft.fields)
    decimal = next(f for f in fields if f["type"] == "decimal")
    decimal["unit"] = "EUR"
    response = api.patch(f"/api/documents/{review_document.pk}/review/", {"version": 1, "fields": fields}, format="json")
    assert response.status_code == 400
    fields = copy.deepcopy(review_draft.fields)
    next(f for f in fields if f["code"] == "role")["value"] = "arbitrary_role"
    review_draft.fields = fields
    review_draft.save()
    draft = api.get(f"/api/documents/{review_document.pk}/review/").data["draft"]
    assert approve(api, review_document, draft).status_code == 400


@pytest.mark.django_db
def test_server_does_not_accept_arbitrary_group_schema(api, review_draft, review_document):
    assert api.post(f"/api/documents/{review_document.pk}/review/groups/", {"version": 1, "group": "participants", "fields": [{"code": "evil"}]}, format="json").status_code == 400


@pytest.mark.django_db
@pytest.mark.parametrize("metadata", ["warnings", "warning_confirmation"])
def test_history_export_checks_warning_and_confirmation_metadata(api, review_draft, review_document, user, metadata):
    data = dict(document=review_document, engine_result=review_draft.engine_result, number=1, draft_version=1, fields=review_draft.fields,
                profile=review_draft.profile, document_name="DANE TESTOWE.pdf", document_checksum=review_document.checksum, author=user)
    data[metadata] = ["DANE\u0001TESTOWE"] if metadata == "warnings" else {"note": "DANE\u0001TESTOWE"}
    revision = ApprovedRevision.objects.create(**data)
    assert api.get(f"/api/revisions/{revision.pk}/export/").status_code == 400


@pytest.mark.django_db(transaction=True)
def test_two_concurrent_group_additions_do_not_duplicate_operation(review_draft, review_document, user):
    from concurrent.futures import ThreadPoolExecutor
    from threading import Barrier

    from django.db import close_old_connections
    from rest_framework.test import APIClient

    barrier = Barrier(2)

    def add():
        close_old_connections()
        api = APIClient()
        api.force_authenticate(user)
        barrier.wait(timeout=5)
        try:
            return api.post(f"/api/documents/{review_document.pk}/review/groups/", {"version": 1, "group": "participants"}, format="json").status_code
        finally:
            close_old_connections()

    with ThreadPoolExecutor(max_workers=2) as pool:
        results = list(pool.map(lambda _: add(), range(2)))
    assert sorted(results) == [200, 409]
    review_draft.refresh_from_db()
    assert len({f["group_id"] for f in review_draft.fields if f["group"] == "participants"}) == 2


def test_common_participant_label_is_not_a_person_name_in_legacy_labelled_profile():
    result = BrokerMotorEngine().extract([PageText(1, "text", "DANE TESTOWE\nWniosek komunikacyjny\nUbezpieczający/Ubezpieczony: Anna Demonstracyjna\n")])
    actual = values(result)
    assert actual["participants.0.role"] == "policyholder,insured"
    assert actual["participants.0.name"] == "Anna Demonstracyjna"
    assert len({f["group_id"] for f in result["fields"] if f["group"] == "participants"}) == 1


def test_manual_rescue_draft_is_in_dashboard_review_count(api, customer, user):
    from documents.models import Document
    from extraction.models import ReviewDraft
    document = Document.objects.create(client=customer, author=user, file="synthetic/manual-rescue.pdf", original_name="DANE TESTOWE ratunek.pdf", mime_type="application/pdf", size=1, checksum="0" * 64)
    ReviewDraft.objects.create(document=document, engine_result=None, profile="broker_motor_application_v1", origin="manual", fields=[])
    response = api.get("/api/dashboard/")
    assert response.status_code == 200
    assert response.data["review_count"] == 1
    assert response.data["review_documents"][0]["id"] == document.pk
