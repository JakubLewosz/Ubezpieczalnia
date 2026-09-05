"""review_export_v0 — kontrolny arkusz, wyłącznie z zatwierdzonego snapshotu."""

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extraction.engine import typed_value
from .text import validate_tree, validate_xlsx_text

PROFILE = "review_export_v0"
NOTICE = "Eksport kontrolny — układ demonstracyjny, do uzgodnienia. DANE TESTOWE."
HEADERS = ["Grupa", "Indeks", "Kod pola", "Nazwa pola", "Wartość", "Typ", "Jednostka", "Strona źródłowa", "Korekta ręczna", "Brak w dokumencie"]


def text_cell(cell, value):
    # Set the underlying XLSX type, including =,+,-,@ or whitespace-prefixed strings.
    # An apostrophe would alter the original value, so do not prepend one.
    validate_xlsx_text(value)
    cell.value = str(value) if value is not None else ""
    cell.data_type = "s"
    cell.number_format = "@"


def build_workbook(revision):
    # Preflight before openpyxl can truncate text or raise an unhandled XML error.
    for name in ["document_name", "document_checksum", "profile"]:
        validate_xlsx_text(getattr(revision, name), name)
    for field in revision.fields:
        for key in ["group", "code", "label", "value", "type", "unit"]:
            validate_xlsx_text(field.get(key), f"{field.get('code', 'pole')}.{key}")
    validate_tree(getattr(revision, "warning_confirmation", {}), "Potwierdzenie ostrzeżeń")
    validate_tree(getattr(revision, "warnings", []), "Ostrzeżenia zatwierdzenia")
    workbook = Workbook()
    info = workbook.active
    info.title = "Informacje"
    rows = [
        ("Ostrzeżenie", NOTICE), ("Profil eksportu", PROFILE),
        ("Dokument", revision.document_name), ("ID dokumentu", str(revision.document_id)),
        ("SHA-256 dokumentu", revision.document_checksum), ("ID zatwierdzonej rewizji", str(revision.pk)),
        ("Numer rewizji", str(revision.number)), ("Profil odczytu", revision.profile),
        ("Zatwierdzono", revision.created_at.isoformat()),
    ]
    # Existing immutable revisions keep their previous worksheet layout.
    if getattr(revision, "warning_confirmation", {}):
        rows.extend([
            ("Pochodzenie szkicu", getattr(revision, "origin", "engine")),
            ("Notatka zatwierdzenia", revision.warning_confirmation.get("note", "")),
        ])
        for warning in getattr(revision, "warnings", []):
            rows.append(("Ostrzeżenie przy zatwierdzeniu", warning.get("message", "") if isinstance(warning, dict) else warning))
    for row_index, pair in enumerate(rows, start=1):
        for col_index, value in enumerate(pair, start=1):
            text_cell(info.cell(row_index, col_index), value)
    info.column_dimensions["A"].width = 29
    info.column_dimensions["B"].width = 100
    info["B1"].alignment = Alignment(wrap_text=True, vertical="top")
    info.row_dimensions[1].height = 35
    sheet = workbook.create_sheet("Dane")
    sheet.append(HEADERS)
    for index, field in enumerate(revision.fields, start=2):
        values = [field["group"], field["index"], field["code"], field["label"], field["value"], field["type"], field["unit"], field["page"], "Tak" if field["manual"] else "Nie", "Tak" if field["absent"] else "Nie"]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(index, col)
            if col == 5 and value is not None:
                parsed = value
                try:
                    if field["type"] != "text":
                        parsed = typed_value(value, field["type"])
                except (ValueError, ArithmeticError):
                    # Świadomie zatwierdzona sprzeczność: zachowaj wierny tekst,
                    # nigdy nie udawaj poprawnej liczby lub daty w XLSX.
                    parsed = None
                if field["type"] != "text" and parsed is None:
                    text_cell(cell, value)
                elif field["type"] == "date":
                    cell.value = date.fromisoformat(parsed)
                    cell.number_format = "yyyy-mm-dd"
                elif field["type"] in {"decimal", "integer"}:
                    cell.value = Decimal(parsed) if field["type"] == "decimal" else int(parsed)
                    cell.number_format = "#,##0.00" if field["type"] == "decimal" else "0"
                else:
                    text_cell(cell, value)
            elif col in {2, 8}:
                cell.value = value
            else:
                text_cell(cell, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for sheet_item in [info, sheet]:
        sheet_item.freeze_panes = "A2"
        for cell in sheet_item[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill("solid", fgColor="245B64")
    for index, width in enumerate([19, 9, 25, 39, 48, 14, 14, 19, 18, 23], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
